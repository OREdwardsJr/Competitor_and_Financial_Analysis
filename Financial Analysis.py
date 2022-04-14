'''
Company names have been redacted with "parent_company" and "child_company"
'''

'''Section 0 - Configurations/Initiations'''
#Import modules and define variables
import requests
import openpyxl
import os
from bs4 import BeautifulSoup
import pandas as pd
import time
import numpy as np
from pandas import ExcelWriter

#Set paths
bb_path = "Absolute path to bb_file"
parent_url = "Home url to jewelry"

#Obtain BB Values
sku_lists = [[]] 
bb_dict = {}
sheet_list = []
a = 0
startTime = time.time()


'''Section 1'''
#Gather Info from BB
for filename in os.listdir(bb_path):
    if filename.endswith('.xlsx'):
        bb_wb = openpyxl.load_workbook(bb_path + filename, data_only = True)
        for sheetname in bb_wb.sheetnames:
            if bb_wb[sheetname].sheet_state == 'visible':
                rotation = filename[:8] + ' ' + sheetname
                sheet_list.append(rotation)            
                bb_dict[rotation] = []
                bb_ws = bb_wb[sheetname]
                #Grab Column Numbers
                for i, col in enumerate(bb_ws.iter_cols(min_row = 5, max_row = 5)):
                    for cell in col:
                        if type(cell.value) == str:
                            if 'sku' in cell.value.lower():
                                sku_col = i + 1
                            if 'product type' in cell.value.lower():
                                prod_col = i+ 1
                            if 'python qty' in cell.value.lower():
                                qty_col = i + 1
                            if 'python cost' in cell.value.lower():
                                cost_col = i + 1
                            if 'python total sell' in cell.value.lower():
                                total_col = i + 1
                            if 'python total child_company cost' in cell.value.lower():
                                child_company_cost_col = i + 1
                            if 'python customer high' in cell.value.lower():
                                cust_high_col = i + 1
                            if 'python customer low' in cell.value.lower():
                                cust_low_col = i + 1
                        i+=1
                #Grab Values
                row_num = 6 
                while True: 
                    if bb_ws.cell(row = row_num, column = sku_col).value != None:
                        if '-' in bb_ws.cell(row = row_num, column = sku_col).value:
                            sku = bb_ws.cell(row = row_num, column = sku_col).value.split('-')[0]
                        else:
                            sku = bb_ws.cell(row = row_num, column = sku_col).value
                        prod_type = bb_ws.cell(row = row_num, column = prod_col).value
                        qty = bb_ws.cell(row = row_num, column = qty_col).value
                        cost = bb_ws.cell(row = row_num, column = cost_col).value
                        total = bb_ws.cell(row = row_num, column = total_col).value
                        child_company_cost = bb_ws.cell(row = row_num, column = child_company_cost_col).value
                        cust_high = bb_ws.cell(row = row_num, column = cust_high_col).value
                        cust_low = bb_ws.cell(row = row_num, column = cust_low_col).value
                        sku_lists[a] = [sku, prod_type, qty, cost, total, 
                                            child_company_cost, cust_high, cust_low]
                        bb_dict[rotation].append(sku_lists[a])
                        sku_lists.append([])
                        a+=1
                    else:
                        break                        
                    j+=1
            bb_list = list(filter(None, sku_lists))

#Performance check of speed in relation to amount of files.
'''
print(len(bb_list))
executionTime = (time.time() - startTime)
print('Execution time in seconds: ' + str(executionTime))
'''

'''Section 2'''
#Scrape Company's website and fill dataframes loop
sku_range = len(bb_dict) 
x = 0
save_path = "Absolute path for saving file"
parent_link = "URL to company website"
frames = {} #Data later to be placed into respective worksheets

while x in range(sku_range):
    startTime = time.time()
    for grab in bb_dict[sheet_list[x]]: 
        sku_link = parent_link + grab[0] 
        page = requests.get(sku_link)
        soup = BeautifulSoup(page.content, 'html.parser')
        run = 0
        while run < 1:
            try:
                strike_div= soup.find("div", class_ = 'strike-price')
                current_div= soup.find("div", class_ = 'current-price')
                strike_price = strike_div.find("span", class_="price-value").text
                current_price = current_div.find("span", class_='price-value').text
            except AttributeError: #Item doesn't exist on company's page.
                strike_price = 'N/A'
                current_price = 'N/A'
            finally: 
                grab.append(strike_price)
                grab.append(current_price)
                run+=1
    try:
        df = pd.DataFrame(bb_dict[sheet_list[x]], columns = ['SKU', 'Product Type', 'Qty', 'Cust Cost/Pc',
                                                             'Total Cust Cost', 'child_company Cost/Pc',
                                                      'Est Cust High', 'Est Cust Low', 'parent_company High', 'parent_company Low'])
    except:
        print(sheet_list[x], ' error') #Catch errors to bb worksheets
        df = pd.DataFrame(bb_dict[sheet_list[x]][:10], columns = ['SKU', 'Product Type', 'Qty', 'Cust Cost/Pc',
                                                             'Total Cust Cost', 'child_company Cost/Pc',
                                                      'Est Cust High', 'Est Cust Low', 'parent_company High', 'parent_company Low'])
    frames[sheet_list[x]] = df    #Save dataframes into a dictionary for later reference
    x+=1

#Save dataframes to Excel
m = 0
w = ExcelWriter('/Users/orlando/Desktop/child_company/Financial Analysis/Analysis.xlsx')
while m in range(len(frames)):
    df = frames[sheet_list[m]]
    df.to_excel(w, sheet_name=sheet_list[m])
    m+=1
w.save()
    
    
'''Section 3'''
#Clean Data
#Set display for max columns
pd.set_option('max_columns', None)

#Make copy of dataframes for debug
frames_copy = {}
for i in range(len(sheet_list)):
    frames_copy[sheet_list[i]] = pd.DataFrame(frames[sheet_list[i]])

#Remove rows with that have products not found on company's website
for i in range(len(sheet_list)):
    filt = (frames_copy[sheet_list[i]]['parent_company High'] == 'N/A')
    frames_copy[sheet_list[i]] = frames_copy[sheet_list[i]][-filt] #Adding '-' in front of filt filters out the information

#Remove rows with #REF in Est Cust High
i = 0
for i in range(len(sheet_list)):
    mask = frames_copy[sheet_list[i]]['Est Cust High'].isin(['#REF!'])
    frames_copy[sheet_list[i]] = frames_copy[sheet_list[i]][-mask]
    i+=1

#Remove $ from High and Low and Convert to Float
for i in range(len(sheet_list)):
    frames_copy[sheet_list[i]]['parent_company High'] = frames_copy[sheet_list[i]]['parent_company High'].str.replace('$','').str.replace('N/A','0').astype(float)
    frames_copy[sheet_list[i]]['parent_company Low'] = frames_copy[sheet_list[i]]['parent_company Low'].str.replace('$','').str.replace('N/A','0').astype(float)


'''Section 4'''
#Peform calculations
i = 0
for i in range(len(sheet_list)):
    qty = frames_copy[sheet_list[i]]['Qty']
    cust_cost_pc = frames_copy[sheet_list[i]]['Cust Cost/Pc']
    total_cust_cost = frames_copy[sheet_list[i]]['Total Cust Cost']
    child_company_cost_pc = frames_copy[sheet_list[i]]['child_company Cost/Pc']
    est_cust_high = frames_copy[sheet_list[i]]['Est Cust High']
    est_cust_low = frames_copy[sheet_list[i]]['Est Cust Low']
    parent_company_high = frames_copy[sheet_list[i]]['parent_company High']
    parent_company_low = frames_copy[sheet_list[i]]['parent_company Low']

    # High Price Diff
    frames_copy[sheet_list[i]]['High Price Diff'] = est_cust_high - parent_company_high
    high_price_diff = frames_copy[sheet_list[i]]['High Price Diff']

    # % High Price Diff
    frames_copy[sheet_list[i]]['% High Price Diff'] = high_price_diff / est_cust_high * 100 
    high_price_perc_diff = frames_copy[sheet_list[i]]['% High Price Diff']

    # % End High Price Share
    frames_copy[sheet_list[i]]['% High Price Share'] = child_company_cost_pc / parent_company_high * 100
    high_price_share = frames_copy[sheet_list[i]]['% High Price Share']

    # Low Price Diff
    frames_copy[sheet_list[i]]['Low Price Diff'] = est_cust_low - parent_company_low
    low_price_diff = frames_copy[sheet_list[i]]['Low Price Diff']

    # % Low Price Diff  
    frames_copy[sheet_list[i]]['% Low Price Diff'] = low_price_diff / est_cust_low * 100 
    low_price_perc_diff = frames_copy[sheet_list[i]]['% Low Price Diff']

    # % End Low Price Share
    frames_copy[sheet_list[i]]['% Low Price Share'] = child_company_cost_pc / parent_company_low* 100
    low_price_share = frames_copy[sheet_list[i]]['% Low Price Share']

    # Earnings/Pc
    frames_copy[sheet_list[i]]['Earnings/Pc'] = (cust_cost_pc - child_company_cost_pc) 
    earnings_per_piece = frames_copy[sheet_list[i]]['Earnings/Pc']

    # Earning Margin/Pc
    frames_copy[sheet_list[i]]['Earnings Margin/Pc'] = earnings_per_piece / child_company_cost_pc
    earnings_margin = frames_copy[sheet_list[i]]['Earnings Margin/Pc']

    # parent_company Earning Per Cost High
    frames_copy[sheet_list[i]]['parent_company Earnings High'] = (parent_company_high - cust_cost_pc) 
    parent_company_earnings_high = frames_copy[sheet_list[i]]['parent_company Earnings High']

    # parent_company Earning Per Cost Low
    frames_copy[sheet_list[i]]['parent_company Earnings Low'] = (parent_company_low - cust_cost_pc)
    parent_company_earnings_low = frames_copy[sheet_list[i]]['parent_company Earnings Low']

    #parent_company Return High
    frames_copy[sheet_list[i]]['parent_company Return High'] = (parent_company_high - cust_cost_pc) / cust_cost_pc
    parent_company_return_high = frames_copy[sheet_list[i]]['parent_company Return High']

    #parent_company Return 
    frames_copy[sheet_list[i]]['parent_company Return Low'] = (parent_company_low - cust_cost_pc) / cust_cost_pc
    parent_company_return_low = frames_copy[sheet_list[i]]['parent_company Return Low']
    
    #parent_company $ Above Cost
    frames_copy[sheet_list[i]]['parent_company $ Above Cost'] = (
        frames_copy[sheet_list[i]]['Cust Cost/Pc']) - (frames_copy[sheet_list[i]]['child_company Cost/Pc']
                                                                                                    )
    #parent_company $ Above Cost Percentage
    frames_copy[sheet_list[i]]['parent_company $ Above Cost %'] = (
        frames_copy[sheet_list[i]]['parent_company $ Above Cost'] / (frames_copy[sheet_list[i]]['child_company Cost/Pc']
    ) * 100)

    i+=1

#Calculate averages
i = 0
for i in range(len(sheet_list)):
    high_price_diff_avg = frames_copy[sheet_list[i]]['High Price Diff'].mean()
    perc_high_price_diff_avg = frames_copy[sheet_list[i]]['% High Price Diff'].mean()
    perc_high_price_share_avg = frames_copy[sheet_list[i]]['% High Price Share'].mean()
    low_price_diff_avg = frames_copy[sheet_list[i]]['Low Price Diff'].mean()
    perc_low_price_diff_avg = frames_copy[sheet_list[i]]['% Low Price Diff'].mean()
    perc_low_price_share_avg = frames_copy[sheet_list[i]]['% Low Price Share'].mean()
    earnings_marginpc_avg = frames_copy[sheet_list[i]]['Earnings Margin/Pc'].mean()
    parent_company_earnings_high_avg = frames_copy[sheet_list[i]]['parent_company Earnings High'].mean()
    parent_company_earnings_low_avg = frames_copy[sheet_list[i]]['parent_company Earnings Low'].mean()
    earnings_pc_avg = frames_copy[sheet_list[i]]['Earnings/Pc'].mean()
    parent_company_return_high_avg = frames_copy[sheet_list[i]]['parent_company Return High'].mean()
    parent_company_return_low_avg = frames_copy[sheet_list[i]]['parent_company Return Low'].mean()
    parent_company_above_cost = frames_copy[sheet_list[i]]['parent_company $ Above Cost'].mean()
    parent_company_above_cost_perc= frames_copy[sheet_list[i]]['parent_company $ Above Cost %'].mean()

    avg_dict = {
                'High Price Diff': [high_price_diff_avg], 
                '% High Price Diff' : [perc_high_price_diff_avg],
                '% High Price Share': [perc_high_price_share_avg], 
                'Low Price Diff': [low_price_diff_avg],
                '% Low Price Diff': [perc_low_price_diff_avg], 
                '% Low Price Share': [perc_low_price_share_avg], 
                'Earnings Margin/Pc': [earnings_marginpc_avg], 
                'parent_company Earnings High': [parent_company_earnings_high_avg],
                'parent_company Earnings Low': [parent_company_earnings_low_avg], 
                'Earnings/Pc': [earnings_pc_avg], 
                'parent_company Return High': [parent_company_return_high_avg], 
                'parent_company Return Low': [parent_company_return_low_avg],
                'parent_company $ Above Cost':[parent_company_above_cost],
                'parent_company $ Above Cost %': [parent_company_above_cost_perc]
                }

    df_avg = pd.DataFrame(avg_dict)
    frames_copy[sheet_list[i]] = frames_copy[sheet_list[i]].append(df_avg, ignore_index = True)


'''Section 5'''
#Clean Data II
#Set Index as Sku, Rename NaN index to 'Average', Drop NaN
i = 0
for i in range(len(sheet_list)):
    frames_copy[sheet_list[i]] = frames_copy[sheet_list[i]].set_index('SKU').rename(index = {np.nan: 'Average'})

#Rearrange Columns
for i in range(len(sheet_list)):
    frames_copy[sheet_list[i]] = frames_copy[sheet_list[i]][['Product Type','Qty', 'Total Cust Cost', 
                               'Cust Cost/Pc','child_company Cost/Pc', 'parent_company $ Above Cost', 'parent_company $ Above Cost %', 
                               'parent_company High','Est Cust High','High Price Diff',  
                               '% High Price Diff','% High Price Share','parent_company Earnings High', 
                                'parent_company Return High', 'parent_company Low', 'Est Cust Low',
                               'Low Price Diff', '% Low Price Diff', '% Low Price Share',
                               'parent_company Earnings Low', 'parent_company Return Low', 'Earnings/Pc','Earnings Margin/Pc' ]]

#Drop NaN in Average 
for i in range(len(sheet_list)):
    frames_copy[sheet_list[i]]['Average'].replace(np.nan, '')

#Fix est cust cost values
grab_list = []
fix_bb = openpyxl.load_workbook('/Users/orlando/Desktop/child_company/Financial Analysis/ROT 64 NOVEMBER BRAND BUY - FINANCE.xlsx', data_only = True)
fix_ws = fix_bb['QA & DEBUT']
i = 0

#Grab Cust High Est values from fix_bb
for grab in fix_ws.iter_cols(min_row = 5, max_row = 5):
    for cell in grab:
        if cell.value == ' TOTAL CUSTOMER UNIT PRICE':
            cust_high_est_col2 = i
    i+=1

for ind in frames_copy[sheet_list[10]].index:
    grab_list.append(ind)
    

#Grab Cust Est High Prices
fix_count = 6
cust_high_est_list = []
for grab in fix_ws.iter_rows(min_row = 6, min_col = 25, max_col = 25):
    for cell2 in grab:
        if type(cell2.value) == str and cell2.value != None:
            if '-' in cell2.value:
                if cell2.value.split('-')[0] in grab_list:
                    cust_high_est_list.append(fix_ws.cell(row = fix_count, column = 134).value)
            else:
                if cell2.value in grab_list:
                    cust_high_est_list.append(fix_ws.cell(row = fix_count, column = 134).value)
        fix_count +=1
        
        
#Replace frames_count[sheet_list[10]] values with cust_high_est_list
for i in range(len(cust_high_est_list)):
    frames_copy[sheet_list[10]]['Est Cust High'][i] = cust_high_est_list[i]


'''Section 6'''
#Save to Excel
'''
The file is saved to Excel with the updated Df. Maybe utilize python to delete rows with 0 or 'N/A'.
Add averages to the bottom.
Plot points of average
'''

#Saved Checkpoint
m = 0
w = ExcelWriter('/Users/orlando/Desktop/child_company/Financial Analysis/Analysis Filter R3.xlsx')
while m in range(len(frames)):
    df = frames_copy[sheet_list[m]]
    df.to_excel(w, sheet_name=sheet_list[m])
    m+=1
    
w.save()

    
# ----------------------- EVERYTHING BELOW IS DEBUGGING ----------------------- #
#FIX ERRORS (Completed)
'''
- Total Cust Cost didn't get pulled appropriately in some df (fixed)
- Est cust High was pulled incorrectly due to incorrect labeling (Sheet_list 10) Rot 64 QA
    - You'll need to fix this by pulling from the Brand Buy somehow (fixed)
- Inf got inserted somehow into [25]. Ref SRA4419. This messes up the mean (Fixed)
- Sheet list 26 has incorrect values in the price diff. IDK how this happened. Happening in others too (Fixed)

Stopped at 25. Finish the rest

01/01/22 - All errors are fixed besides point 2
'''
frames_copy[sheet_list[21]]['Est Cust High'][0] - float(frames_copy[sheet_list[21]]['parent_company High'][0].replace('$','')) 

#print(sheet_list[20])


#Debug V1.0
#Sheet_list was getting the same variable name thus causing append function within parent_company to append to it twice

#Scrape parent_company Individually
'''
Scrape using one bb at a time. It will take too long to do it all at once. Rotations are contained within
bb_dict. Sheet_list is a list of rotation names. It's how you can identify what rot goes with each bb_dict
'''
startTime = time.time()
parent_link = 'https://www.parent_company.com/product/'

for grab in bb_dict[sheet_list[24]]: #Change grab
    sku_link = parent_link + grab[0] #Change grab
    page = requests.get(sku_link)
    soup = BeautifulSoup(page.content, 'html.parser')
    k = 0
    j = 0
#    strike_div= soup.find("div", class_ = 'strike-price')
#    current_div= soup.find("div", class_ = 'current-price')
    try:
        strike_div= soup.find("div", class_ = 'strike-price')
        current_div= soup.find("div", class_ = 'current-price')
        k+=1
        if k >1:
            print(k, 'k', grab)
        strike_price = strike_div.find("span", class_="price-value").text
        current_price = current_div.find("span", class_='price-value').text
        j+=1
        if j > 1:
            print(j, 'j', grab)
    except AttributeError:
        strike_price = 'N/A'
        current_price = 'N/A'
    finally: 
        grab.append(strike_price)  #Change grab
        grab.append(current_price)  #Change grab
    df = pd.DataFrame(bb_dict[sheet_list[24]], columns = ['SKU', 'Product Type', 'Qty', 'Cust Cost/Pc',
                                                             'Total Cust Cost', 'child_company Cost/Pc',
                                                      'Est Cust High', 'Est Cust Low', 'parent_company High', 'parent_company Low'])


for sheet in sheet_list:
    print(sheet)

executionTime = (time.time() - startTime)
print('Execution time in seconds: ' + str(executionTime))